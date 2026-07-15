import csv
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl is not installed. Please install it using: pip install openpyxl")
    exit(1)

def main():
    csv_path = Path("all_test_cases.csv")
    excel_path = Path("real_life_test_cases.xlsx")
    filtered_csv_path = Path("real_life_test_cases.csv")

    if not csv_path.exists():
        print(f"Error: {csv_path} does not exist.")
        return

    # Create a new workbook and active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Passed Test Cases"

    # Set grid lines visible
    ws.views.sheetView[0].showGridLines = True

    # Styling helpers
    font_family = "Segoe UI"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Dark Navy Blue
    
    cell_font = Font(name=font_family, size=10)
    passed_font = Font(name=font_family, size=10, bold=True, color="276A3C")
    passed_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid") # Light Green

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    passed_rows = []

    with open(csv_path, mode="r", encoding="utf-8") as f:
        reader = csv.reader(f)
        headers = next(reader)
        
        # Write headers
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = thin_border
        
        # Filter and write rows
        row_num = 2
        for row in reader:
            if not row:
                continue
            
            # Check status (column index 7) and tested (column index 6)
            tested = row[6] if len(row) > 6 else ""
            status = row[7] if len(row) > 7 else ""
            
            # ONLY include test cases that are Tested: Yes and Status: Passed
            if tested == "Yes" and status == "Passed":
                passed_rows.append(row)
                ws.append(row)
                
                # Format row cells
                for col_idx in range(1, len(row) + 1):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.font = cell_font
                    cell.border = thin_border
                    
                    if col_idx in [1, 7, 8]:
                        cell.alignment = align_center
                        cell.font = Font(name=font_family, size=10, bold=True)
                    else:
                        cell.alignment = align_left
                    
                    # Highlight passed fields
                    if col_idx in [7, 8]:
                        cell.font = passed_font
                        cell.fill = passed_fill
                
                row_num += 1

    # Save filtered CSV
    with open(filtered_csv_path, mode="w", newline="", encoding="utf-8") as f_out:
        writer = csv.writer(f_out)
        writer.writerow(headers)
        writer.writerows(passed_rows)

    # Adjust row heights and column widths if rows exist
    if row_num > 2:
        ws.row_dimensions[1].height = 28
        for r in range(2, row_num):
            ws.row_dimensions[r].height = 22

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 40)

    # Save Excel workbook
    wb.save(excel_path)
    print(f"Successfully generated formatted Excel sheet (ONLY Passed cases): {excel_path}")
    print(f"Successfully generated clean CSV (ONLY Passed cases): {filtered_csv_path}")

if __name__ == "__main__":
    main()
